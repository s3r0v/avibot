import requests
import openpyxl
import time
from random import randint
import os
import re
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

def parse(links, name, available_nums):
    REGION_INFO = 'region_info'
    CATEGORIES_INFO = 'categories_info'
    AVITO_MAIN = 'avito_main'
    #available_nums = int(balance/11)
    sessids = file_to_array('sessid.txt')
    sessid = str(sessids[-1])
    numbers_flag = 0

    cookie = {"name":"sessid", "value":sessid}

    options = webdriver.FirefoxOptions()
    options.set_preference("general.useragent.override", "Mozilla/95.0.2 (iPhone; U; CPU like Mac OS X; en) AppleWebKit/420+ (KHTML, like Gecko) Version/3.0 Mobile/1A543 Safari/419.3")
    options.set_preference('permissions.default.image', 2)
    options.set_preference('dom.ipc.plugins.enabled.libflashplayer.so', 'false')
    options.add_argument('--no-sandbox')
    options.headless = True
    driver = webdriver.Firefox(options=options)
    driver.get("https://m.avito.ru")
    driver.add_cookie(cookie)
    nums = []
    links = links[:available_nums+1]

    for link in links:
        try:
            driver.get(link)
            #WebDriverWait(driver,10).until(EC.element_to_be_clickable(By.XPATH, "//button[@class='mav-vpko6w']")).click()
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, "//button[@class='mav-vpko6w']"))).click()
            time.sleep(1)
            nums.append(driver.find_element_by_xpath("//span[@data-marker='phone-popup/phone-number']").text)

        except Exception as e:
            print(e)
            nums.append("-")

    for num in nums:
        if num != "-":
            numbers_flag+=1

    #delete_last_line('sessid.txt')
    numbers_flag = array_to_file(nums, name, numbers_flag)
    return numbers_flag
    

def handle_excel(file):
    wb = openpyxl.load_workbook(filename = file)
    ws = wb.active
    links = []

    columns = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

    for col in ws.iter_cols():
        for cell in col:
            if "Ссылка" in str(cell.value):
                column=cell.column
    for cell in ws[columns[column-1]]:
        if cell.value!=None:
            links.append(cell.value)
    return links[1:]

def file_to_array(file):
    arr = []
    with open(file) as f:
        while True:
            line = f.readline()
            if not line:
                break
            arr.append(line.strip())
    return arr

def delete_last_line(file):
    with open(file, "r+", encoding = "utf-8") as file:

        file.seek(0, os.SEEK_END)

        pos = file.tell() - 1

        while pos > 0 and file.read(1) != "\n":
            pos -= 1
            file.seek(pos, os.SEEK_SET)

        if pos > 0:
            file.seek(pos, os.SEEK_SET)
            file.truncate()

def array_to_file(nums, name, numbers_flag):
    wb = openpyxl.load_workbook(name)
    sheets = wb.sheetnames
    sheet = wb[sheets[0]]
    f = 0

    for row in sheet:
        if f == 0:
            numbers = row[-1].offset(column=1)
            numbers.value = "Номера"
        else:
            numbers = row[-1].offset(column=1)
            try:
                numbers.value = nums[f-1]
            except Exception:
                numbers.value = "-"
        f+=1
    wb.save(name)

    return numbers_flag

def connect_proxy(proxy_flag):
    proxy = file_to_array("proxies.txt")[proxy_flag].split(":")
    proxies = {'http':f'http://{proxy[0]}:{proxy[1]}'}
    print(proxies)
    return proxies

def delete_promocode(text):
    with open('promocodes.txt') as f:
        lines = f.readlines()

    str = text
    pattern = re.compile(re.escape(str))
    with open('promocodes.txt', 'w') as f:
        for line in lines:
            result = pattern.search(line)
            if result is None:
                f.write(line)
